#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
prepare_mannino_stats.py

Parse SINTEF MSS-Adjusts Surgery Data and create clean duration statistics
that can be used to build Mannino-style OR scheduling instances.

Recommended use:
    python prepare_mannino_stats.py ^
      --input mssadjustssurgerydata.xls ^
      --output-dir mannino_prepared ^
      --elective-only ^
      --min-duration 5 ^
      --max-duration 600

Dependencies:
    For .xls:  pip install xlrd
    For .xlsx: pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple


HEADER_ALIASES = {
    "Year": ["Year"],
    "Month": ["Month"],
    "week": ["week", "Week"],
    "Surgery Team": ["Surgery Team", "Team", "Specialty"],
    "Actual Surgery TIME": ["Actual Surgery TIME", "Actual Surgery Time", "Duration", "Surgery duration"],
    "Emergency": ["Emergency"],
}


def _match_columns(headers: List[str]) -> Dict[str, int]:
    result = {}
    normalized = {str(h).strip(): i for i, h in enumerate(headers)}
    for canonical, aliases in HEADER_ALIASES.items():
        for a in aliases:
            if a in normalized:
                result[canonical] = normalized[a]
                break
    missing = [k for k in HEADER_ALIASES if k not in result]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Headers found: {headers}")
    return result


def _read_xls(path: Path) -> Tuple[List[str], List[List[Any]]]:
    try:
        import xlrd
    except ImportError as e:
        raise ImportError("Reading .xls requires xlrd. Install with: pip install xlrd") from e

    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
    rows = []
    for r in range(1, sheet.nrows):
        rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    return headers, rows


def _read_xlsx(path: Path) -> Tuple[List[str], List[List[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("Reading .xlsx requires openpyxl. Install with: pip install openpyxl") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    rows = [list(r) for r in rows_iter]
    return headers, rows


def read_workbook(path: Path) -> Tuple[List[str], List[List[Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported file type: {suffix}. Use .xls or .xlsx")


def percentile(values: List[float], q: float) -> float:
    if not values:
        return float("nan")
    xs = sorted(values)
    pos = (len(xs) - 1) * q
    lo = math.floor(pos)
    hi = math.ceil(pos)
    if lo == hi:
        return xs[int(pos)]
    return xs[lo] * (hi - pos) + xs[hi] * (pos - lo)


def summarize(values: List[float]) -> Dict[str, float]:
    return {
        "n": len(values),
        "mean": statistics.mean(values) if values else float("nan"),
        "sd": statistics.stdev(values) if len(values) > 1 else 0.0,
        "min": min(values) if values else float("nan"),
        "q10": percentile(values, 0.10),
        "q25": percentile(values, 0.25),
        "median": percentile(values, 0.50),
        "q75": percentile(values, 0.75),
        "q90": percentile(values, 0.90),
        "max": max(values) if values else float("nan"),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Path to mssadjustssurgerydata.xls or converted .xlsx")
    parser.add_argument("--output-dir", default="mannino_prepared")
    parser.add_argument("--elective-only", action="store_true", help="Keep Emergency == No only")
    parser.add_argument("--min-duration", type=float, default=5.0, help="Remove tiny/invalid durations below this number of minutes")
    parser.add_argument("--max-duration", type=float, default=600.0, help="Remove extreme durations above this number of minutes")
    args = parser.parse_args()

    path = Path(args.input)
    outdir = Path(args.output_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    headers, rows = read_workbook(path)
    cols = _match_columns([str(h).strip() for h in headers])

    clean = []
    for row in rows:
        team = row[cols["Surgery Team"]]
        dur = row[cols["Actual Surgery TIME"]]
        emerg = row[cols["Emergency"]]
        if team is None or str(team).strip() == "":
            continue
        try:
            dur = float(dur)
        except Exception:
            continue
        if dur < args.min_duration or dur > args.max_duration:
            continue
        emerg_str = str(emerg).strip()
        if args.elective_only and emerg_str.lower() != "no":
            continue
        clean.append({
            "year": row[cols["Year"]],
            "month": row[cols["Month"]],
            "week": row[cols["week"]],
            "team": str(team).strip(),
            "duration_min": dur,
            "emergency": emerg_str,
        })

    if not clean:
        raise RuntimeError("No records left after cleaning. Check filters and input columns.")

    # Case-level output.
    case_csv = outdir / "mannino_case_pool.csv"
    with open(case_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["year", "month", "week", "team", "duration_min", "emergency"])
        writer.writeheader()
        writer.writerows(clean)

    # Summary by team.
    by_team: Dict[str, List[float]] = {}
    for r in clean:
        by_team.setdefault(r["team"], []).append(float(r["duration_min"]))

    total = sum(len(v) for v in by_team.values())
    summary_rows = []
    summary_json = {
        "source_file": str(path),
        "elective_only": bool(args.elective_only),
        "min_duration": args.min_duration,
        "max_duration": args.max_duration,
        "total_records": total,
        "teams": {},
    }

    for team, vals in sorted(by_team.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        s = summarize(vals)
        pct = 100.0 * len(vals) / total
        row = {
            "team": team,
            "count": len(vals),
            "percentage": pct,
            "duration_mean": s["mean"],
            "duration_sd": s["sd"],
            "duration_min": s["min"],
            "duration_q10": s["q10"],
            "duration_q25": s["q25"],
            "duration_median": s["median"],
            "duration_q75": s["q75"],
            "duration_q90": s["q90"],
            "duration_max": s["max"],
        }
        summary_rows.append(row)
        summary_json["teams"][team] = row

    summary_csv = outdir / "mannino_duration_summary.csv"
    with open(summary_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
        writer.writeheader()
        writer.writerows(summary_rows)

    summary_path = outdir / "mannino_duration_stats.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary_json, f, indent=2)

    print(f"Clean records: {total}")
    print(f"Wrote: {case_csv}")
    print(f"Wrote: {summary_csv}")
    print(f"Wrote: {summary_path}")


if __name__ == "__main__":
    main()
